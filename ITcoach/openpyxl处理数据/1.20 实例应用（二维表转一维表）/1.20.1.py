# import openpyxl
# # wb=openpyxl.load_workbook('业绩表.xlsx')
# # ws=wb.worksheets[0]
# # if '转换表' in wb.sheetnames:
# #     wb.remove(wb['转换表'])
# # nws=wb.create_sheet('转换表');nws.append(['姓名','月份','业绩'])
# # rng1=ws['a'][1:]
# # rng2=ws.iter_rows(min_col=2,min_row=2)
# # for name,row in list(zip(rng1,rng2)):
# #     for x,y in zip(ws[1][1:],row):
# #         nws.append([name.value,x.value,y.value])
# # wb.save('业绩表.xlsx')




import openpyxl
wb = openpyxl.load_workbook("业绩表.xlsx")
ws = wb.active
print(ws.row_values(1))
# l = [[col.value for col in row] for row in ws.rows]
# # print(l)
# month = l[0][1:]
# # print(month)
#
# nws = wb.create_sheet("一维表")
# nws.append(["姓名","月份","业绩"])
# name = [n[0] for n in l[1:]]
# # print(name)
# data = [n[1:] for n in l[1:]]
# # print(data)
# for n, row in zip(name, data):
#     for m, c in zip(month, row):
#         nws.append([n, m, c])
# wb.save("业绩表.xlsx")
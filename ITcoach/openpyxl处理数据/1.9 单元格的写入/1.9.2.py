# import openpyxl
# wb=openpyxl.Workbook()
# ws=wb.active
# ws.title='九九表'
# for x in range(1,10):
#     for y in range(1,x+1):
#         ws.cell(x,y,'%d×%d=%d'%(y,x,x*y))
# wb.save('九九表.xlsx')

import openpyxl

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "九九乘法表"
for i in range(1,10):
    for j in range(1,i+1):
        ws.cell(i,j,"%d×%d＝%d" %(j,i,i*j))
wb.save("九九乘法表.xlsx")





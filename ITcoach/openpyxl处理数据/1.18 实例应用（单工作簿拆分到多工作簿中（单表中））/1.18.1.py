# import openpyxl
# wb=openpyxl.load_workbook('工资表.xlsx',data_only=True)
# ws=wb.worksheets[0]
# rngs=list(ws.values)
# d={}
# for row in rngs[1:]:
#     if row[2] in d.keys():
#         d[row[2]]+=[row]
#     else:
#         d.update({row[2]:[row]})
# for k,v in d.items():
#     nwb=openpyxl.Workbook()
#     nws=nwb.active;nws.title=k
#     nws.append(rngs[0])
#     for r in v:
#         nws.append(r)
#     wb.save('各部门工资表/'+k+'.xlsx')

import openpyxl
import os

wb = openpyxl.load_workbook("工资表.xlsx",data_only=True)
ws = wb.active
caption = list(ws.values)[0]
department_index = caption.index("部门")
# print(department_index)
department_dict = {}
for row in list(ws.values)[1:]:
    if row[department_index] in department_dict.keys():
        department_dict[row[department_index]] += [row]
    else:
        department_dict[row[department_index]] = [row]
print(department_dict)

# 当前文件的路径
current_dir = os.path.dirname(os.path.abspath(__file__))
# 创建文件夹
folder_name = "各部门工资表2"
# 创建文件夹的地址
folder_address = os.path.join(current_dir, folder_name)
# 判断文件夹是否存在，不存在就创建
if not os.path.exists(folder_address):
    os.mkdir(folder_address)

for k,v in department_dict.items():
    nwb = openpyxl.Workbook()
    nws = nwb.active
    nws.title = k
    nws.append(caption)
    for row in v:
        nws.append(row)
    nwb.save(folder_name + '/' + k + '.xlsx')

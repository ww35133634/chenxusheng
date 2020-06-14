# import openpyxl
# wb=openpyxl.load_workbook('各年业绩表.xlsx')
# nwb=openpyxl.Workbook()
# nws=nwb.active
# nws.append(['年份','月份','金额'])
# for ws in wb.worksheets:
#     l=[[ws.title]+[c.value for c in row] for row in ws.rows][1:-1]
#     for v in l:
#         nws.append(v)
# nwb.save('合并1.xlsx')

import openpyxl

wb = openpyxl.load_workbook("各年业绩表.xlsx")
nws = wb.create_sheet("汇总表",0)
nws.append(["年份","月份","业绩"])
for ws in wb.worksheets[:-1]:
    print([[ws.title] +[c.value for c in row] for row in list(ws.rows)][1:-1])
    l = [[ws.title] +[c.value for c in row] for row in list(ws.rows)][1:-1]
    for v in l:
        nws.append(v)
wb.save("各年业绩表01.xlsx")









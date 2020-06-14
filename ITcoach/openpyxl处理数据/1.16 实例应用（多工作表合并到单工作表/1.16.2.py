# import openpyxl
# wb=openpyxl.load_workbook('各年业绩表.xlsx')
# nwb=openpyxl.Workbook()
# nws=nwb.active
# nws.append(['年份','月份','金额'])
# for ws in wb.worksheets:
#     l=list(ws.values)[1:-1]
#     print(l)
#     for v in l:
#         nws.append((ws.title,)+v)
# nwb.save('合并2.xlsx')

import openpyxl
wb = openpyxl.load_workbook("各年业绩表.xlsx")
nwb = openpyxl.Workbook()
nws = nwb.active
nws.title = "业绩汇总"
nws.append(['年份','月份','金额'])
for ws in wb.worksheets:
    print([ws.title])
    l = list(ws.values)[1:-1]
    for v in l:
        print(v)
        print((ws.title,) + v)
        nws.append((ws.title,) + v)
nwb.save("业绩汇总.xlsx")




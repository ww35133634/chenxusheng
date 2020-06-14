# import openpyxl
# wb=openpyxl.load_workbook('test.xlsx')
# ws=wb.worksheets[0]
# rngs=ws[2:ws.max_row]
# nws=wb.create_sheet('结果')
# nws.append(['姓名','总分'])
# for row in rngs:
#     nws.append([row[0].value,sum([c.value for c in row][1:])])
# wb.save('test1.xlsx')

import openpyxl

wb = openpyxl.load_workbook("test.xlsx")
ws = wb.active
nws = wb.create_sheet("结果表")
nws.append(["姓名","总分"])
rngs = list(ws.iter_rows())
for row in rngs[1:]:
    print(row[0].value,sum([col.value for col in row[1:]]))
    nws.append([row[0].value,sum([col.value for col in row[1:]])])
wb.save("test2.xlsx")


















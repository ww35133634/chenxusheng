# import openpyxl
# wb=openpyxl.load_workbook('test.xlsx')
# ws=wb.active
#
# f=[sum(l)/len(l) for l in list(zip(*list(ws.values)[1:]))[1:]]
# n=[c.value for c in ws['1']][1:]
# print(['%s-%.2f'%c for c in list(zip(n,f))])
#
# print(['%s-%.2f'%(l[0],sum(l[1:])/len(l[1:])) for l in list(zip(*list(ws.values)))[1:]])

import openpyxl
import numpy as np
wb =openpyxl.load_workbook("test.xlsx")
ws = wb.worksheets[0]
print(["%.2f-%s" %(np.average(l[1:]),l[0]) for l in list(zip(*list(ws.values)))[1:]])

print([np.average(l[1:]) for l in list(zip(*list(ws.values)))[1:]])

course = [l[0] for l in list(zip(*list(ws.values)))[1:]]
avg_result = [np.average(l[1:]) for l in list(zip(*list(ws.values)))[1:]]

for c in zip(course,avg_result):
    print("%s-%.2f"%c)


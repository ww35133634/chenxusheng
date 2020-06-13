# import openpyxl
# wb=openpyxl.load_workbook('demo.xlsx',data_only=True)
# ws=wb.active
# # print([[c.value for c in row] for row in ws['a1:d3']])
# print(list(ws.values)[1:4])

import openpyxl

wb = openpyxl.load_workbook("demo.xlsx",data_only=True)
ws = wb.active
# print(list(ws.values)[1:])
# print([sum(v) for v in [c[2:][:-1] for c in list(ws.values)[1:]]])
print([[v.value for v in r] for r in ws["a1:d3"]])


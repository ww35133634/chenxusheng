# import openpyxl
# wb=openpyxl.load_workbook('demo.xlsx')
# ws=wb.active
# rngs=ws.iter_rows(min_row=2,min_col=2)
# for row in rngs:
#         sm=sum([c.value for c in row][0:4])
#         if sm>=300:
#             row[-1].value='优秀'
# wb.save('demo2.xlsx')
#
# import openpyxl
# wb = openpyxl.load_workbook("demo.xlsx")
# ws = wb.active
# nws = wb.create_sheet("转换后成绩表")
# nws.append(list(ws.values)[0])
# data = [list(row) for row in list(ws.values)[1:]]
# print(data)
# result = [sum(c[1:-2]) for c in data]
# print(result)
# for index,rw in enumerate(data):
#     rw[-2] = result[index]
#     if result[index] >= 300:
#         rw[-1] = "优秀"
#     nws.append(rw)
# wb.save("demo-1.xlsx")

import openpyxl

wb = openpyxl.load_workbook("demo.xlsx")
ws = wb.active
rngs = ws.iter_rows(min_row=ws.min_row + 1, min_col=ws.min_column + 1)

for row in rngs:
    result = sum([c.value for c in row][0:-2])
    row[-2].value = result
    if result >= 300:
        row[-1].value = "优秀"
wb.save("demo-2.xlsx")





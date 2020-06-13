# import openpyxl
# wb=openpyxl.load_workbook('test.xlsx')
# ws=wb.worksheets[0]
# total=[sum([c.value for c in row]) for row in ws.iter_rows(min_col=2,min_row=2)]
# name=[c.value for c in ws['a']][1:]
# print(list(zip(name,total)))

import openpyxl
wb = openpyxl.load_workbook("test.xlsx")
ws = wb.active

# print([sum([col.value for col in row]) for row in ws.iter_rows(min_row=2,min_col=2)])

total = [sum([col.value for col in row]) for row in ws.iter_rows(min_row=2,min_col=2)]
name = [name.value for name in list(ws["a"])[1:]]
# print(name)
print(list(zip(name,total)))









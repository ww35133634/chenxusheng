# import openpyxl
# wb=openpyxl.load_workbook('test.xlsx')
# ws=wb.worksheets[0]
# for row in ws.iter_rows(min_row=36,min_col=2,max_col=4,max_row=40):
#     print([c.value for c in row])

import openpyxl

wb = openpyxl.load_workbook("test.xlsx")
ws = wb.active
print([[c.value for c in row] for row in ws.iter_rows()])

print([[c for c in row] for row in ws.values])








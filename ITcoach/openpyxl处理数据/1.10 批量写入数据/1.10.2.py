# import openpyxl
# wb=openpyxl.load_workbook('demo.xlsx')
# ws=wb.active
# # for r in [['%d*%d=%d'%(y,x,x*y) for y in range(1,x+1)] for x in range(1,10)]:
# #     ws.append(r)
# # ws.delete_rows(1)
# # wb.save('demo.xlsx')
# for row in ws['a1:c6']:
#     for c in row:
#         c.value=1
# wb.save('demo.xlsx')

import openpyxl
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "列表推导九九表"

for r in [["%d×%d＝%d"%(j,i,i*j) for j in range(1,i+1)] for i in range(1, 10)]:
    ws.append(r)
wb.save("九九表.xlsx")
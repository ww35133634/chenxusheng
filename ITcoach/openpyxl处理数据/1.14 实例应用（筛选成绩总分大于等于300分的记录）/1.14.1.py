# import openpyxl
# wb=openpyxl.load_workbook('test.xlsx')
# ws=wb.active
# rngs=list(ws.rows)
# nws=wb.create_sheet('筛选结果')
# nws.append([c.value for c in rngs[0]]+['总分'])
# for row in rngs[1:]:
#     l=[c.value for c in row]
#     l+=[sum(l[1:])]
#     if l[-1]>=300:
#         nws.append(l)
# wb.save('test1.xlsx')

import openpyxl
wb = openpyxl.load_workbook("test.xlsx")
ws = wb.active
rngs = list(ws.rows)
nws = wb.create_sheet("结果表")
nws.append([c.value for c in rngs[0]]+["总分"])
for row in rngs[1:]:
    l = [col.value for col in row]
    print(l)
    total = [sum([col.value for col in row][1:])]
    print(total)
    l += total
    print(l)
    if l[-1] >= 300:
        nws.append(l)
wb.save("test01.xlsx")













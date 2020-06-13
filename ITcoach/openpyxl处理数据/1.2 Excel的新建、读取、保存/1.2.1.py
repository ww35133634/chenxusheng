# import openpyxl
# wb=openpyxl.Workbook()
# wb.save('我的工作簿.xlsx')

import openpyxl

# wb = openpyxl.Workbook()
# wb.save("newworkbook.xlsx")

# wb = openpyxl.load_workbook("newworkbook.xlsx")
# ws = wb.active
# print(ws.title)

for i in range(1,13):
    wb = openpyxl.Workbook()
    wb.save("%d月" %i)
# import openpyxl
# # # wb=openpyxl.Workbook()
# # # wb.create_sheet()
# # # wb.create_sheet()
# # # wb.create_sheet()
# # # wb.save('demo1.xlsx')
# #
# # wb=openpyxl.load_workbook('demo2.xlsx')
# # wb.create_sheet('工资表',2)
# # wb.save('demo2.xlsx')

import openpyxl
wb = openpyxl.load_workbook("de-1.xlsx")
print(wb.sheetnames)
wb.remove(wb["1-2月"])
wb.save("de-2.xlsx")







# import openpyxl
# wb=openpyxl.load_workbook('各班成绩表.xlsx')
# ws=wb.active
# rngs=list(ws.values)
# d={}
# for row in rngs[1:]:
#     if row[0] in d.keys():
#         d[row[0]]+=[row]
#     else:
#         d[row[0]]=[row]
# nwb=openpyxl.Workbook()
# for k,v in sorted(d.items()):
#     nws=nwb.create_sheet(k)
#     nws.append(rngs[0])
#     for r in v:
#         nws.append(r)
# nwb.remove(nwb.worksheets[0])
# nwb.save('拆分.xlsx')

import openpyxl

wb = openpyxl.load_workbook("各班成绩表.xlsx")
ws = wb.active
grade = dict()
for row in list(ws.values)[1:]:
    if row[0] in grade.keys():
        grade[row[0]] += [row]
    else:
        grade[row[0]] = [row]
nwb = openpyxl.Workbook()

for k,v in sorted(grade.items()):
    nws = nwb.create_sheet(k)
    nws.append(list(ws.values)[0])
    for r in v:
        nws.append(r)
nwb.remove(nwb.worksheets[0])
nwb.save("拆分到表.xlsx")























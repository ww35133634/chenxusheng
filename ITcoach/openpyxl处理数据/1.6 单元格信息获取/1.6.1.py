# import openpyxl
# wb=openpyxl.load_workbook('demo.xlsx')
# ws=wb.worksheets[0]
# print(ws['b1'].value)
# print(ws.cell(1,2).value)
# print(openpyxl.load_workbook('demo.xlsx').worksheets[0]['b1'].value)

import openpyxl

# wb = openpyxl.load_workbook("demo.xlsx")
# # ws = wb.active
# ws = wb.worksheets[0]
# # ws = wb["Sheet1"]
# # print(ws)
# print(ws['b1'].value)
# print(ws.cell(1,2).value)


wb = openpyxl.load_workbook("各年业绩表.xlsx")

total = 0

for ws in wb.worksheets:
    total += ws["b14"].value
print(total)
print(sum([s["b14"].value for s in wb.worksheets]))

# import xlrd,xlwt
# wb=xlrd.open_workbook('业绩表.xls')
# nwb=xlwt.Workbook(encoding='utf-8')
# nws=nwb.add_sheet('结果')
# #---------------------------------
# def counter(list):
#     return len([x for x in list if x>=100])
# #---------------------------------
# n=0
# for s in wb.sheets():
#     n+=1
#     nws.write(n,0,s.name)
#     l=s.col_values(1)[1:]
#     nws.write(n,1,list(map(counter,[l]))[0])
# nws.write(0,0,'月份')
# nws.write(0,1,'计数')
# nwb.save('统计结果1.xls')

import openpyxl

def read_excel():
    wb = openpyxl.load_workbook("业绩表.xlsx")
    ws = wb.worksheets
    return ws
def  get_data():
    ws = read_excel()
    results = []
    results.append(["月份", "计数"])
    for s in ws:
        result = []
        i = 0
        for r in list(s.rows)[1:]:
            if r[1].value >= 100:
                i += 1
        result.append(s.title)
        result.append(i)
        results.append(result)
    return results

def save_data():
    results = get_data()
    nwb = openpyxl.Workbook()
    nws = nwb.active
    nws.title = "结果"
    print(results)
    for row in results:
        nws.append(row)
    nwb.save("统计结果01.xlsx")

if __name__ == '__main__':
    save_data()


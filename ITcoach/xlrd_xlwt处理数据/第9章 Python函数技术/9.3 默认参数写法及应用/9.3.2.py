# import xlrd
# from xlutils.copy import copy
# #------------------------------------------
# def text_pos(str,delimiter,position=0):
#     n=0;l=[-1]
#     while n<str.count(delimiter):
#         l.append(str.index(delimiter,l[-1]+1))
#         n+=1
#     l=l[1:]
#     if position==0:
#         l=l[0]
#     elif position==1:
#         l=l[-1]
#     elif position==2:
#         l=l[:]
#     return l
# def mid(text,start,num):
#     txt=text[start:start+num]
#     return txt
# #------------------------------------------
# wb=xlrd.open_workbook('产品信息表.xls')
# ws=wb.sheet_by_name('产品表')
# nwb=copy(wb);nws=nwb.get_sheet('产品表')
# n=0
# for c in ws.col_values(1)[1:]:
#     n+=1
#     ll=[mid('，'+c,x+1,6) for x in text_pos('，'+c,'，',2)]
#     nws.write(n,2,'/'.join(ll))
# nwb.save('产品信息表.xls')

import openpyxl

def get_text_code(one_text):
    one_text_code = []
    split_text = one_text.split("，")
    for one_split_text in split_text:
        one_text01 = one_split_text.split("/")[0]
        one_text_code.append(one_text01)
    return one_text_code

wb = openpyxl.load_workbook("产品信息表.xlsx")
ws = wb.active
ws.cell(1,4).value = "抽取信息"
all_text_list = [v[1] for v in list(ws.values)[1:]]
r = 2
for one_text in all_text_list:
    text = get_text_code(one_text)
    one_row = "、".join(text)
    ws.cell(r, 4, one_row)
    r += 1
wb.save("产品信息表01.xlsx")





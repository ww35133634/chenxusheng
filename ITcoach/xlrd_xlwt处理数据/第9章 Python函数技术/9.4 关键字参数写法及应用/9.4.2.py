# import xlrd
# from xlutils.copy import copy
# #------------------------------------------------
# def agg(list,delimiter,num=1,subtotal=sum):
#     l = [int(x.split(delimiter)[num]) for x in list]
#     return subtotal(l)
# #------------------------------------------------
#
# wb=xlrd.open_workbook('2018年业绩表.xls')
# ws=wb.sheet_by_name('2018业绩表')
# nwb=copy(wb);nws=nwb.get_sheet('2018业绩表')
# r=0
# while r<ws.nrows-1:
#     r+=1
#     l=ws.row_values(r)[1:-1]
#     nws.write(r,6,agg(delimiter=',',list=l))
# nwb.save('2018年业绩表.xls')

import openpyxl

def get_subtotal(list,delimiter):
    subtotal = sum([int(x.split(delimiter)[1]) for x in list])
    return subtotal
    # all_one_num = []
    # for one in list:
    #     one_num = int(one.split(delimiter)[1])
    #     all_one_num.append(one_num)
    # return sum(all_one_num)

wb = openpyxl.load_workbook("2018年业绩表.xlsx")
ws = wb.worksheets[0]
print(ws.max_column)
r = 2
for row in ws.iter_rows(min_row=2, min_col=1):
    one_row = [r.value for r in row][1:-1]
    subtotal = get_subtotal(one_row, ",")
    ws.cell(r,ws.max_column).value = subtotal
    r += 1
wb.save("2018年业绩表01.xlsx")




























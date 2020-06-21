# import xlrd
# from xlutils.copy import copy
# #-------------------------------------
# def age(id,dit):
#     l=[id[x:y] for x,y in ((6,10),(10,12),(12,14))]
#     date=dit.join(l)
#     return date
# #-------------------------------------
# wb=xlrd.open_workbook('用户信息.xls')
# ws=wb.sheet_by_name('信息表')
# nwb=copy(wb);nws=nwb.get_sheet('信息表')
# r=0
# while r<ws.nrows-1:
#     r+=1
#     nws.write(r,2,age(ws.cell_value(r,1),'-'))
# nwb.save('用户信息.xls')

import openpyxl

def get_birth(one_id_num):
    birth_num = [one_id_num[x:y] for x,y in ((6,10),(10,12),(12,14))]
    birthday = "-".join(birth_num)
    return birthday


wb = openpyxl.load_workbook("用户信息.xlsx")
ws = wb["信息表"]
all_id_num = [birth[1] for birth in list(ws.values)[1:]]
r = 2
c = 4
ws.cell(1,c).value = "出生日期"
for one_id_num in all_id_num:
    birthday = get_birth(one_id_num)
    ws.cell(r,c).value = birthday
    r += 1
wb.save("用户信息01.xlsx")










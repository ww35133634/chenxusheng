import openpyxl
from day05_list.class_random import class_random


class ReadExcel:
    def __init__(self, path: str, sheet: str, infos: list, flag: bool = True):
        self.path = path
        self.sheet = sheet
        self.infos = infos  # dict类型，需要用的Key
        self.flag = flag  # 如果是True,表示数据有表头，

        # 读取数据保存 集合 样式
        self.rows_list01 = []  # 格式1 [[],[],...]
        self.rows_list02 = []  # 格式1 [{},{},...]

    def read_excel(self):
        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook[self.sheet]
        for index, row in enumerate(sheet.rows):
            one_row_list = []
            one_row_dict = {}
            if index == 0 and self.flag:
                continue
            for col_index, col_value in enumerate(row):
                if col_index == 3:
                    dt = col_value.value.strftime("%Y-%m-%d")
                    one_row_list.append(dt)
                    one_row_dict[self.infos[col_index]] = dt
                else:
                    one_row_list.append(col_value.value)
                    one_row_dict[self.infos[col_index]] = col_value.value
            self.rows_list01.append(one_row_list)
            self.rows_list02.append(one_row_dict)


class WriteExcel:
    def __init__(self, path: str, sheet: str, infos: list, flag: bool = True):
        self.path = path
        self.sheet = sheet
        self.infos = infos
        self.flag = flag

    def write_excel(self, data: list):
        """ data数据格式是 [[],[],...] """
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = self.sheet
        if len(self.infos) > 0 and self.flag:
            data.insert(0, self.infos)

        for index, row in enumerate(data):
            for col_index, col_value in enumerate(row):
                sheet.cell(row=index + 1, column=col_index + 1, value=col_value)
        workbook.save(self.path)





# if __name__ == '__main__':
    # #     """读取excel"""
    # #     path = r"F:\python\ITcoach\day07_oprate_excel\student01.xlsx"
    # #     sheet = "student"
    # #     infos = ['sno', 'name', 'gender', 'birthday', 'mobile', 'email', 'address']
    # #     obj01 = ReadExcel(path, sheet, infos, True)
    # #     obj01.read_excel()
    # #     print(obj01.rows_list01)
    # #     print(obj01.rows_list02)

if __name__ == '__main__':
    """写入excel"""
    path = r"F:\python\ITcoach\day07_oprate_excel\student01.xlsx"
    sheet = "student"
    infos = ['name', 'chinese', 'math', 'english', 'physics', 'chemistry']
    data_list = []
    data01 = class_random.Random_name(10)
    for index, name in enumerate(data01.name):
        temp = []
        temp.append(name)
        data02 = class_random.Random_result(50, 100, 5)   #生成 1条 5个不重复的 随机数据
        temp.extend(data02.one_distinc_result)
        data_list.append(temp)
    print(data_list)
    obj02 = WriteExcel(path,sheet,infos,True)
    obj02.write_excel(data_list)


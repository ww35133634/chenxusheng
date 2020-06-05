import openpyxl

# 读取
# path = r"C:\Users\CXS\Desktop\公司文件\ERP数据\供应商预计进货表.xlsx"
# sheet = "Sheet1"
# info = ['供应商编号', '供应商简称', '预交货日', '品号', '供应商品号', '品名', '规格', '采购数量', '已交数量', '未交数量', '待验中量', '单位', '小单位', '结束', '币种', '汇率', '单价', '已交金额', '未交金额', '采购单号', '采购日期', '交货仓库', '采购人员', '工单号', '备    注', '采购包装数量', '已交包装数量', '包装单位', '未交包装数量', '请购单号', '赠备品数量', '赠备品已交数量']
# workbook =openpyxl.load_workbook(path)
# sheet = workbook[sheet]
# rows_list = []
# for index, row in enumerate(sheet.rows):
#     one_row_list = []
#     # one_row_dit = {}
#     for col_index,col_value in enumerate(row):
#         # one_row_dit[info[col_index]] = col_value.value
#         one_row_list.append(col_value.value)
#     rows_list.append(one_row_list)
#     # rows_list.append(one_row_dit)
# # print(rows_list)
#
#
# # 写入
# path = r"C:\Users\CXS\Desktop\公司文件\ERP数据\供应商预计进货表02.xlsx"
# workbook = openpyxl.Workbook()
# sheet = workbook.active
# sheet.title = "sheet1"
# for index,row in enumerate(rows_list):
#     for col_index,col_value in enumerate(row):
#         sheet.cell(row=index + 1,column=col_index + 1,value=col_value)
# workbook.save(path)
# 新建表

class ReadExcel:
    def __init__(self,path,sheet,flag:bool=False):
        self.path = path
        self.sheet = sheet
        self.flag = flag         #标题，默认False无标题
        self.rows_list01 = []    #格式1 [[],[],...]
        self.rows_list02 = []    #格式1 [{},{},...]
    def read_excel(self):
        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook[self.sheet]
        if self.flag == False:
            for index,row in enumerate(sheet.rows):
                one_row_list = []
                for col_index,col_value in enumerate(row):
                    one_row_list.append(col_value.value)
                self.rows_list01.append(one_row_list)



if __name__ == '__main__':
    path = r"C:\Users\CXS\Desktop\公司文件\ERP数据\供应商预计进货表.xlsx"
    sheet = "Sheet1"
    info = ['供应商编号', '供应商简称', '预交货日', '品号', '供应商品号', '品名', '规格', '采购数量', '已交数量', '未交数量', '待验中量', '单位', '小单位', '结束',
            '币种', '汇率', '单价', '已交金额', '未交金额', '采购单号', '采购日期', '交货仓库', '采购人员', '工单号', '备    注', '采购包装数量', '已交包装数量',
            '包装单位', '未交包装数量', '请购单号', '赠备品数量', '赠备品已交数量']
    obj01 = ReadExcel(path,sheet)
    obj01.read_excel(2,info)
    print(obj01.rows_dict_list)
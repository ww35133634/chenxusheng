from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
import os
import time
import openpyxl
def writeExcel():
    # 获取文件路径
    excelPath = os.path.join(os.getcwd(), 'ExcelData')
    print
    "****"
    print
    excelPath
    # 定义文件名称
    # invalid mode ('wb') or filename: 'Excel2017-09-21_20:15:57.xlsx'  这种方式明明文件，会提示保存失败，无效的文件名。
    # nameTime = time.strftime('%Y-%m-%d_%H:%M:%S')
    nameTime = time.strftime('%Y-%m-%d_%H-%M-%S')
    excelName = 'Excel' + nameTime + '.xlsx'
    ExcelFullName = os.path.join(excelPath, excelName)
    print
    ExcelFullName

    wb = Workbook()

    ws = wb.active

    tableTitle = ['userName', 'Phone', 'age', 'Remark']

    # 维护表头
    #    if row < 1 or column < 1:
    #     raise ValueError("Row or column values must be at least 1")
    # 如上，openpyxl 的首行、首列 是 （1,1）而不是（0,0），如果坐标输入含有小于1的值，提示 ：Row or column values must be at least 1，即最小值为1.
    for col in range(len(tableTitle)):
        c = col + 1
        ws.cell(row=1, column=c).value = tableTitle[col]

    # 数据表基本信息
    tableValues = [['张学友', 15201062100, 18, '测试数据！'], ['李雷', 15201062598, 19, '测试数据！'],
                   ['Marry', 15201062191, 28, '测试数据！']]

    for row in range(len(tableValues)):
        ws.append(tableValues[row])
    # wb.save(ExcelFullName)
    wb.save(filename=ExcelFullName)
    return ExcelFullName


def readExcel(path):
    workbook = openpyxl.load_workbook(path)
    # # workbook = openpyxl.load_workbook(filename=ExcelFullName)
    # sheet = workbook.active  #获取当前活跃的worksheet,默认就是第一个worksheet
    # print(sheet.title)  #输出表格名称
    sheets = workbook.get_sheet_names()   #获取所有表格(worksheet)的名字
    print(sheets)
    sheet_first = str(sheets[0])  #第一个表格的名称
    print(sheet_first)
    sheet = workbook.get_sheet_by_name(sheet_first)  #获取特定的worksheet

    # 获取表格所有行和列，两者都是可迭代的
    rows = sheet.rows
    columns = sheet.columns

    # 迭代所有的行
    for row in rows:
        line = [col.value for col in row]
        print(line)

    # 通过坐标读取值
    print(sheet['A1'].value)  # A表示列,1表示行
    print(sheet.cell(row=1, column=1).value)


if __name__ == '__main__':
    path = r"C:\Users\CXS\Desktop\cj.xlsx"
    readExcel(path)

    # ExcelFullName = writeExcel()
    # readExcel(ExcelFullName)